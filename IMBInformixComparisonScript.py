# coding=utf-8
from SharedFunctions import *
import glob
from Server import Server
from openpyxl.workbook.workbook import Workbook


# origin_folder = input("Directory : ")
origin_folder = "tbschema_2017_05_23"
origin_folder += "/*.txt"
origin_folder_glob = glob.glob(origin_folder)

input_folder = origin_folder[:-6] + "_inputs"
if not os.path.exists(input_folder):
    os.makedirs(input_folder)

servers_dictionnary = {}

for origin_file in origin_folder_glob:
    origin_file_name = origin_file[len(origin_folder[:-5])]
    input_file_name = locals()['origin_file_name'][:-4] + "input.txt"
    input_file = origin_folder + "_inputs" + input_file_name
    prepare_tbschema_file(origin_file, input_file)
    server_name = origin_file_name.split('-')[1]

    if server_name not in list(servers_dictionnary.keys()):
        servers_dictionnary[server_name] = Server(input_file)
    else:
        servers_dictionnary[server_name].add_database_to_dictionnary(input_file)


excel_file = Workbook()
ws = excel_file.active
ws.title = "Summary"

ws.cell(row=1, column=1, value='ACTION')
ws.cell(row=1, column=2, value='COMPL.')
ws.cell(row=1, column=3, value='DB')
ws.cell(row=1, column=4, value='TABLE')

tbl_cat_dict = {}
idx_cat_dict = {}
rvk_cat_dict = {}
grt_cat_dict = {}
server_column = 5
line_number = 2

for server in list(servers_dictionnary.keys()):
    line_number_found = 2
    srv = servers_dictionnary.get(server)
    ws.cell(row=1, column=server_column, value=srv.server_name)

    for database in list(srv.dictionnary_databases.keys()):
        db = srv.dictionnary_databases.get(database)

        #####
        # Tables comparison
        #####
        for table in list(db.tables_dictionnary.keys()):
            tbl = db.tables_dictionnary.get(table)

            # I'm looking for the row where the cols A, B, C & D are already filled
            found = False

            for row in ws.iter_rows(min_col=1, max_col=4):
                if row[0].internal_value == 'TABLE' and row[2].internal_value == db.database_name and row[3].internal_value == tbl.table_name:
                    line_number_found = row[0].row
                    found = True
                    break

            # If not found, I append a new row with my info
            tbl_cat_dict_key = ""
            if not found:
                ws.cell(row=line_number, column=1, value='TABLE')
                ws.cell(row=line_number, column=3, value=db.database_name)
                ws.cell(row=line_number, column=4, value=tbl.table_name)
                tbl_cat_dict_key = db.database_name + "-" + tbl.table_name
                tbl_cat_dict[tbl_cat_dict_key] = tbl
                ws.cell(row=line_number, column=server_column, value=list(tbl_cat_dict.keys()).index(tbl_cat_dict_key))
                line_number += 1

            # If found, I add my info into the good cell
            else:
                # I check into my tbl_cat_dict if I've the same table
                same_cat = False
                for category in list(tbl_cat_dict.keys()):
                    cat = tbl_cat_dict.get(category)
                    if cat.__eq__(tbl):
                        same_cat = True
                        break

                if not same_cat:
                    tbl_cat_dict_key = db.database_name + "-" + tbl.table_name
                    tbl_cat_dict[tbl_cat_dict_key] = tbl

                ws.cell(row=line_number_found, column=server_column, value=list(tbl_cat_dict.keys()).index(tbl_cat_dict_key))

        #####
        # Indexes comparison
        #####
        for index in list(db.indexes_dictionnary.keys()):
            idx = db.indexes_dictionnary.get(index)

            # I'm looking for the row where the cols A, B, C & D are already filled
            found = False

            for row in ws.iter_rows(min_col=1, max_col=4):
                if row[0].internal_value == 'INDEX' and row[1].internal_value == idx.unique_constraint and row[2].internal_value == db.database_name and row[3].internal_value == idx.index_name:
                    line_number_found = row[0].row
                    found = True
                    break

            # If not found, I append a new row with my info
            idx_cat_dict_key = ""
            if not found:
                ws.cell(row=line_number, column=1, value='INDEX')
                ws.cell(row=line_number, column=2, value=idx.unique_constraint)
                ws.cell(row=line_number, column=3, value=db.database_name)
                ws.cell(row=line_number, column=4, value=idx.index_name)
                idx_cat_dict_key = db.database_name + "-" + idx.unique_constraint + "-" + idx.index_name
                idx_cat_dict[idx_cat_dict_key] = idx
                ws.cell(row=line_number, column=server_column, value=list(idx_cat_dict.keys()).index(idx_cat_dict_key))
                line_number += 1

            # If found, I add my info into the good cell
            else:
                # I check into my tbl_cat_dict if I've the same table
                same_cat = False
                for category in list(idx_cat_dict.keys()):
                    cat = idx_cat_dict.get(category)
                    if cat.__eq__(idx):
                        same_cat = True
                        break

                if not same_cat:
                    idx_cat_dict_key = db.database_name + "-" + idx.unique_constraint + "-" + idx.index_name
                    idx_cat_dict[idx_cat_dict_key] = idx

                ws.cell(row=line_number_found, column=server_column, value=list(idx_cat_dict.keys()).index(idx_cat_dict_key))

        #####
        # Revokes comparison
        #####
        for revoke in list(db.revokes_dictionnary.keys()):
            rvk = db.revokes_dictionnary.get(revoke)

            # I'm looking for the row where the cols A, B, C & D are already filled
            found = False
            for row in ws.iter_rows(min_col=1, max_col=4):
                if row[0].internal_value == 'REVOKE' and row[1].internal_value == rvk.privilege_revoked and row[2].internal_value == db.database_name and row[3].internal_value == rvk.table_revoke:
                    line_number_found = row[0].row
                    found = True
                    break

            # If not found, I append a new row with my info
            rvk_cat_dict_key = ""
            if not found:
                ws.cell(row=line_number, column=1, value='REVOKE')
                ws.cell(row=line_number, column=2, value=rvk.privilege_revoked)
                ws.cell(row=line_number, column=3, value=db.database_name)
                ws.cell(row=line_number, column=4, value=rvk.table_revoke)
                rvk_cat_dict_key = db.database_name + "-" + rvk.privilege_revoked + "-" + rvk.table_revoke
                rvk_cat_dict[rvk_cat_dict_key] = rvk
                ws.cell(row=line_number, column=server_column, value=list(rvk_cat_dict.keys()).index(rvk_cat_dict_key))
                line_number += 1

            # If found, I add my info into the good cell
            else:
                # I check into my tbl_cat_dict if I've the same table
                same_cat = False
                for category in list(rvk_cat_dict.keys()):
                    cat = rvk_cat_dict.get(category)
                    if cat.__eq__(rvk):
                        same_cat = True
                        break

                if not same_cat:
                    rvk_cat_dict_key = db.database_name + "-" + rvk.privilege_revoked + "-" + rvk.table_revoke
                    rvk_cat_dict[rvk_cat_dict_key] = rvk

                ws.cell(row=line_number_found, column=server_column, value=list(rvk_cat_dict.keys()).index(rvk_cat_dict_key))

        #####
        # Grants comparison
        #####
        for grant in list(db.grants_dictionnary.keys()):
            grt = db.grants_dictionnary.get(grant)

            # I'm looking for the row where the cols A, B, C & D are already filled
            found = False
            for row in ws.iter_rows(min_col=1, max_col=4):
                if row[0].internal_value == 'GRANT' and row[1].internal_value == grt.privilege_granted and row[2].internal_value == db.database_name and row[3].internal_value == grt.table_grant:
                    line_number_found = row[0].row
                    found = True
                    break

            # If not found, I append a new row with my info
            grt_cat_dict_key = ""
            if not found:
                ws.cell(row=line_number, column=1, value='GRANT')
                ws.cell(row=line_number, column=2, value=grt.privilege_granted)
                ws.cell(row=line_number, column=3, value=db.database_name)
                ws.cell(row=line_number, column=4, value=grt.table_grant)
                grt_cat_dict_key = db.database_name + "-" + grt.privilege_granted + "-" + grt.table_grant
                grt_cat_dict[grt_cat_dict_key] = grt
                ws.cell(row=line_number, column=server_column, value=list(grt_cat_dict.keys()).index(grt_cat_dict_key))
                line_number += 1

                # If found, I add my info into the good cell
            else:
                # I check into my tbl_cat_dict if I've the same table
                same_cat = False
                for category in list(grt_cat_dict.keys()):
                    cat = grt_cat_dict.get(category)
                    if cat.__eq__(grt):
                        same_cat = True
                        break

                if not same_cat:
                    grt_cat_dict_key = db.database_name + "-" + grt.privilege_granted + "-" + grt.table_grant
                    grt_cat_dict[grt_cat_dict_key] = grt

                ws.cell(row=line_number_found, column=server_column, value=list(grt_cat_dict.keys()).index(grt_cat_dict_key))

    server_column += 1

# Calcul de l'homogénéité, manipulation manuelle à faire car je n'ai pas trouvé comment faire des formules matricielles avec OpenPyXL
ws.cell(row=1, column=server_column, value='HOMOGÉNÉITÉ')
ws.cell(row=2, column=server_column, value='{=NB.SI(E2:AL2;INDEX(E2:AL2; EQUIV(FAUX; ESTVIDE(E2:AL2);0)))/COLONNES(E2:AL2)}')

excel_file.save('resultat.xlsx')